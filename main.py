import os.path
from datetime import date

import numpy
import pandas
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows

import StockData
import graphGeneration

companiesFile = "webAdressCompanies.csv"

url_bourse = "https://www.boursorama.com/cours/"
excel_file = "D:/Python_Projects/WebScrapping/stockData.xlsx"
stockList = []
currentVarList = []
weekVarList = []
monthVarList = []
month3VarList = []
month6VarList = []
yearVarList = []
year3VarList = []
year5VarList = []
year10VarList = []


# Read froml CSV into a dictionary.
def readCSV(filename):
    data_dict = {}
    csv_data = pandas.read_csv(filename, sep=',', header=None)
    csv_data_ = csv_data
    csv_data = numpy.array([x.split(';') for x in csv_data_[0]])
    for data in csv_data:
        data_dict[data[0]] = data[1]
    return data_dict


# Function to insert DataFrame below the current date cell
def insert_dataframe(ws, start_row, start_col, df, index_col):
    for r in dataframe_to_rows(df, index=False, header=False):
        ws.append(r)
    if ws["A15"] is not None:
        ws.move_range("A15:D28", rows=-13, cols=4 * index_col)
    for cell in ws[start_row][start_col:start_col + len(df.columns)]:
        cell.alignment = Alignment(horizontal='center')


# Check if output file exists to load it or to create it.
def outputFile_create(filename, stocks_dict):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        companies_list = list(stocks_dict.keys())
        companies_list.sort()
        for key in companies_list:
            if key in wb.sheetnames:
                # print(key, " already present")
                continue
            else:
                # print(key, " newly added")
                wb.create_sheet(key)  # Create a worksheet for newly added company.
    else:
        wb = Workbook()
        for key in stocks_dict.keys():
            wb.create_sheet(key)  # Create a worksheet per company.
        del wb['Sheet']  # Remove worksheet created by default.
        wb.save(filename)
    return wb


def main():
    if os.path.exists(companiesFile):
        stocks_dict = readCSV(companiesFile)
        # Check if output file exists to load it or to create it.
        wb = outputFile_create(excel_file, stocks_dict)

        for key in stocks_dict.keys():
            url_data = url_bourse + stocks_dict[key]
            stock_data = StockData.get_stock_data(url_data)
            stockList.append(key)
            currentVarList.append(stock_data.values[0][1])
            weekVarList.append(stock_data.values[1][1])
            monthVarList.append(stock_data.values[2][1])
            month3VarList.append(stock_data.values[3][1])
            month6VarList.append(stock_data.values[4][1])
            yearVarList.append(stock_data.values[5][1])
            year3VarList.append(stock_data.values[6][1])
            year5VarList.append(stock_data.values[7][1])
            year10VarList.append(stock_data.values[8][1])

            for ws in wb.worksheets:
                if ws.title != key:
                    continue
                else:
                    # Max cols in empty Excel file is 16384. Find the first cell empty and merge and fill it.
                    index_col = 0  # Used as index to shift data after first day recovering data.
                    for i in range(1, 16384, 4):
                        cell = ws.cell(1, i)
                        if cell.value is None:
                            cell.value = date.today().strftime("%d-%m-%Y")
                            ws.merge_cells(None, 1, i, 1, i + 3)
                            cell.alignment = Alignment(horizontal="center")
                            cell.font = Font(bold=True)
                            index_col = index_col + 1
                            break
                        else:
                            # print("Hello ", cell.value)  # Print date of already used cell.
                            continue
                    # Get data in the Excel file.
                    insert_dataframe(ws, 2, 1, stock_data, index_col)
                    wb.save(excel_file)

        graphGeneration.plot_yearVariation(stockList, currentVarList, weekVarList, monthVarList, month3VarList,
                                           month6VarList, yearVarList, year3VarList, year5VarList,
                                           year10VarList)
    else:
        print("No companies to get data from...")
        print("End of the program...")


if __name__ == "__main__":
    main()
